"""Shared column schema + sheet builders for the Design-C rating instruments.

Design C — within-teacher staged disclosure:
  P1  盲评        own diagnosis (no AI) + confidence
  P2a 看结论后    reveal AI diagnosis CONCLUSION only -> re-rate own + confidence
  P2b 再评        reasoning arm: reveal AI reasoning -> re-rate own + confidence
                  control arm:  no new info -> re-rate own + confidence
  P3  评AI质量    reveal question + student answer -> groundedness/question/ICAP/...

Used by gen_teacher_packets.py (per-teacher packets, reasoning gated by arm) and
build_rating_workbook_v2.py (all-551 master, reasoning always shown).
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

FONT = "Arial"
FILLS = {
    "id": PatternFill("solid", fgColor="DDE6F2"), "ctx": PatternFill("solid", fgColor="EEF3FA"),
    "p1": PatternFill("solid", fgColor="E7F4E4"), "p2a": PatternFill("solid", fgColor="CDE7C8"),
    "p2b": PatternFill("solid", fgColor="A9D6A0"), "q": PatternFill("solid", fgColor="FFF6D5"),
    "meta": PatternFill("solid", fgColor="F0F0F0"),
}
HEADCOL = {"id": "1F3864", "ctx": "1F3864", "p1": "2E7D32", "p2a": "1B5E20",
           "p2b": "0E3B0E", "q": "B8860B", "meta": "555555"}
thin = Side(style="thin", color="C9C9C9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
DROPDOWN_KINDS = {"S15", "DIAG_GND", "CONSIST", "ELICIT", "YESNO", "ATTITUDE", "ICAP", "CORRECT", "FAB"}

DV = {"S15": '"1,2,3,4,5"', "DIAG_GND": '"全部有据,个别臆造,多数臆造"', "CONSIST": '"一致,部分一致,矛盾"',
      "ELICIT": '"L1回忆,L2解释,L3论证,L4迁移"', "YESNO": '"是,否"', "ATTITUDE": '"真诚尝试,敷衍,空答或乱码"',
      "ICAP": '"被动,主动,建构,互动"', "CORRECT": '"正确,部分正确,错误"', "FAB": '"有,无"'}

# (header, round-key | special | None, width, dropdown-kind, fill-group)
# specials: "__reasoning__" (gated by arm), "__teacher__", "__arm__"
COLS = [
    ("round_id", "round_id", 11, None, "id"),
    ("学生ID", "student_id", 7, None, "ctx"),
    ("学科", "subject", 7, None, "ctx"),
    ("课题", "topic", 15, None, "ctx"),
    ("轮次", "round", 5, None, "ctx"),
    ("本课学习目标", "objectives_text", 32, None, "ctx"),
    ("学生反思（输入）", "reflection_text", 28, None, "ctx"),
    ("上一轮·AI问题", "prev_q", 18, None, "ctx"),
    ("上一轮·学生答", "prev_a", 14, None, "ctx"),
    # ---- P1 盲评 ----
    ("【P1】盲诊_掌握度(3/2/1/0)", None, 15, None, "p1"),
    ("【P1】我会怎么问", None, 18, None, "p1"),
    ("【P1】信心", None, 7, "S15", "p1"),
    # ---- AI conclusion reveal (input for P2a) ----
    ("AI诊断·已掌握", "ai_mastered_text", 24, None, "ctx"),
    ("AI诊断·未掌握", "ai_not_mastered_text", 24, None, "ctx"),
    ("AI针对目标(理由)", "ai_targets_text", 22, None, "ctx"),
    # ---- P2a 看结论后重评 ----
    ("【P2a】看结论后_掌握度(3/2/1/0)", None, 16, None, "p2a"),
    ("【P2a】信心", None, 7, "S15", "p2a"),
    # ---- P2b input: reasoning (reasoning arm only) ----
    ("AI推理过程〔仅推理臂〕", "__reasoning__", 36, None, "ctx"),
    # ---- P2b 再评 (reasoning arm: after reasoning; control arm: no new info) ----
    ("【P2b】再评_掌握度(3/2/1/0)", None, 16, None, "p2b"),
    ("【P2b】信心", None, 7, "S15", "p2b"),
    ("【P2b】改判说明(可选)", None, 18, None, "p2b"),
    # ---- P3 评AI质量 ----
    ("【Q】诊断groundedness", None, 13, "DIAG_GND", "q"),
    ("【Q】臆造说明", None, 15, None, "q"),
    ("【Q】与盲诊(P1)一致性", None, 12, "CONSIST", "q"),
    ("【Q】诊断有用性", None, 8, "S15", "q"),
    ("AI追问", "ai_question", 28, None, "ctx"),
    ("【Q】追问相关性", None, 8, "S15", "q"),
    ("【Q】对准薄弱点", None, 9, "S15", "q"),
    ("【Q】引出力", None, 9, "ELICIT", "q"),
    ("【Q】适切", None, 7, "S15", "q"),
    ("【Q】会用吗", None, 8, "YESNO", "q"),
    ("【Q】会用原因", None, 14, None, "q"),
    ("学生回答", "student_answer", 22, None, "ctx"),
    ("【Q】作答态度", None, 12, "ATTITUDE", "q"),
    ("【Q】ICAP", None, 10, "ICAP", "q"),
    ("【Q】深度", None, 7, "S15", "q"),
    ("【Q】正确性", None, 11, "CORRECT", "q"),
    ("【Q】整轮推进", None, 9, "S15", "q"),
    ("【Q】匹配度", None, 8, "S15", "q"),
    # ---- meta ----
    ("评分老师", "__teacher__", 9, None, "meta"),
    ("臂", "__arm__", 10, None, "meta"),
    ("备注", None, 14, None, "meta"),
]

RCOLS = [
    ("round_id", "round_id", 11, None, "id"), ("学科", "subject", 7, None, "ctx"),
    ("课题", "topic", 14, None, "ctx"), ("轮次", "round", 5, None, "ctx"),
    ("学生反思（输入）", "reflection_text", 28, None, "ctx"),
    ("AI诊断·已掌握", "ai_mastered_text", 24, None, "ctx"),
    ("AI诊断·未掌握", "ai_not_mastered_text", 24, None, "ctx"),
    ("AI追问", "ai_question", 26, None, "ctx"), ("AI推理过程（reasoning）", "reasoning", 52, None, "ctx"),
    ("⑥a推理逻辑", None, 9, "S15", "p2b"), ("⑥b faithfulness", None, 11, "S15", "p2b"),
    ("⑥c有无捏造", None, 10, "FAB", "p2b"), ("评分老师", "__teacher__", 9, None, "meta"),
    ("臂", "__arm__", 9, None, "meta"), ("备注", None, 14, None, "meta"),
]

CTRL_PLACEHOLDER = "〔控制臂：本轮不提供推理，请在不看新信息的情况下再评一次〕"


def build_sheet(ws, cols, rounds, *, teacher=None, arm_map=None,
                reasoning_mode="by_arm", dvmap=None, row_height=104):
    """reasoning_mode: 'by_arm' (packets) shows reasoning only for 推理臂 rounds;
    'always' (master) always shows it."""
    dvmap = dvmap if dvmap is not None else DV
    ws.sheet_view.showGridLines = False
    for ci, (hdr, key, width, dv, fg) in enumerate(cols, start=1):
        c = ws.cell(1, ci, hdr)
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=HEADCOL[fg]); c.alignment = CTR; c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 48
    for ri, r in enumerate(rounds, start=2):
        arm = arm_map.get(r["round_id"]) if arm_map else None
        for ci, (hdr, key, width, dv, fg) in enumerate(cols, start=1):
            c = ws.cell(ri, ci); c.border = BORDER; c.font = Font(name=FONT, size=10)
            if key == "__reasoning__":
                show = reasoning_mode == "always" or arm == "推理臂"
                c.value = r["reasoning"] if show else CTRL_PLACEHOLDER
                c.fill = FILLS["ctx"]; c.alignment = WRAP
            elif key == "__teacher__":
                c.value = teacher or ""; c.fill = FILLS["meta"]; c.alignment = CTR
            elif key == "__arm__":
                c.value = arm or ""; c.fill = FILLS["meta"]; c.alignment = CTR
            elif key is not None:
                c.value = r.get(key, ""); c.fill = FILLS["id"] if fg == "id" else FILLS[fg]
                c.alignment = CTR if fg == "id" else WRAP
            else:
                c.fill = FILLS[fg]
                c.alignment = CTR if dv in DROPDOWN_KINDS else WRAP
        ws.row_dimensions[ri].height = row_height
    ws.freeze_panes = "B2"
    nrow = len(rounds) + 1
    for kind, formula in dvmap.items():
        d = DataValidation(type="list", formula1=formula, allow_blank=True, showErrorMessage=False)
        ws.add_data_validation(d)
        for ci, (hdr, key, width, k, fg) in enumerate(cols, start=1):
            if k == kind:
                L = get_column_letter(ci); d.add(f"{L}2:{L}{nrow}")


def build_legend(ws):
    """Write the 使用说明 sheet (phases + arm + scale legend)."""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 72
    HEAD = PatternFill("solid", fgColor="1F3864")

    def row(r, a, b="", c="", head=False, bold=False, fill=None):
        ws.cell(r, 1, a); ws.cell(r, 2, b); ws.cell(r, 3, c)
        for col in (1, 2, 3):
            cc = ws.cell(r, col)
            cc.font = Font(name=FONT, bold=bold or head, size=14 if head else 11,
                           color="FFFFFF" if head else "000000")
            cc.alignment = WRAP
            if head:
                cc.fill = HEAD
            elif fill:
                cc.fill = fill

    r = 1
    ws.cell(r, 1, "研道街 · 四阶段教师评分（方案 C）").font = Font(name=FONT, bold=True, size=16, color="1F3864")
    r += 2
    row(r, "四个阶段（务必按顺序，不回退）", head=True); r += 1
    for ln in [
        "P1 盲评（浅绿）：看不到 AI，给出自己的逐目标掌握度、我会怎么问、信心(1–5)。",
        "P2a 看结论后（中绿）：揭示 AI 的诊断【结论】（已掌握/未掌握），据此【重新】给出自己的掌握度 + 信心。",
        "P2b 再评（深绿）：推理臂 → 揭示 AI【推理过程】后再评一次 + 信心；控制臂 → 不给新信息，直接再评一次 + 信心。",
        "P3 评 AI 质量（黄）：再揭示 AI 追问与学生回答，对 AI 诊断/追问/学生表现打分。",
        "最后：到「推理审计」表评 AI 推理本身（逻辑/faithfulness/有无捏造）。",
    ]:
        row(r, "", "", ln); r += 1
    r += 1
    row(r, "⚠ 关于「臂」", head=True); r += 1
    row(r, "", "", "每行「臂」列标明 推理臂 / 控制臂。推理臂行在 P2b 看 AI 推理；控制臂行 P2b 不给新信息、只是再评一次（用于量化“单纯再看一遍”的漂移基线）。两臂已配平，请按本行实际显示作答。"); r += 2

    row(r, "评分图例", head=True); r += 1
    ws.cell(r, 1, "列"); ws.cell(r, 2, "取值"); ws.cell(r, 3, "含义")
    for col in (1, 2, 3):
        ws.cell(r, col).font = Font(name=FONT, bold=True, color="FFFFFF")
        ws.cell(r, col).fill = PatternFill("solid", fgColor="2E5496")
        ws.cell(r, col).alignment = WRAP
    r += 1
    legend = [
        ("P1/P2a/P2b 掌握度", "3/2/1/0", "3已掌握 2部分 1未掌握 0无法判断。逐目标写，如“目标1=2; 目标2=0”。", "p1"),
        ("P1/P2a/P2b 信心", "1–5", "对当前这个判断有多确信。5=非常确信。", "p2a"),
        ("P2b 改判说明", "文本(可选)", "若改了，为什么改；尽量写依据 AI 的哪句话。", "p2b"),
        ("Q 诊断groundedness", "全部有据/个别臆造/多数臆造", "AI 每条掌握判断是否有反思原文支持。", "q"),
        ("Q 与盲诊(P1)一致性", "一致/部分一致/矛盾", "AI 诊断与你 P1 盲评的异同。", "q"),
        ("Q 诊断有用性", "1–5", "5精准可教 3偏泛 1没用/误导。", "q"),
        ("Q 追问相关性 / 对准薄弱点", "1–5", "扣住本轮反思 / 戳中薄弱处的程度。", "q"),
        ("Q 引出力", "L1/L2/L3/L4", "L1回忆 L2解释 L3论证 L4迁移；L3/L4 才算“答辩式”。", "q"),
        ("Q 适切 / 会用吗", "1–5 / 是否", "难度适切；你会不会真对这孩子问这句。", "q"),
        ("Q 作答态度", "真诚尝试/敷衍/空答或乱码", "学生是否认真作答。", "q"),
        ("Q ICAP", "被动/主动/建构/互动", "被动复述；主动操作无解释；建构自己的解释；互动在追问上修正深化。", "q"),
        ("Q 深度 / 整轮推进 / 匹配度", "1–5", "回答展开程度；本轮是否逼出新思考；追问与水平匹配度。", "q"),
        ("Q 正确性", "正确/部分正确/错误", "背景项，可选。", "q"),
        ("推理审计 6a/6b/6c", "1–5 / 有无", "AI 推理逻辑、是否真支撑结论(faithfulness)、有无捏造。", "p2b"),
    ]
    for name, vals, desc, fg in legend:
        row(r, name, vals, desc, fill=FILLS[fg])
        ws.cell(r, 1).font = Font(name=FONT, bold=True, size=11)
        r += 1
