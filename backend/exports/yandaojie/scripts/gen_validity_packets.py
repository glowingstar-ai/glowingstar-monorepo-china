"""Generate per-subject teacher packets for the DEFENSE-DEPTH ANNOTATION study.

This is the load-bearing reliability instrument for the validity-gap thesis:
teachers label, per defense episode, (1) the MC-view (would answer-only scoring
pass this objective?) and (2) the defense depth (sound / right-label-wrong-
mechanism / hollow-or-guess / no-understanding). MC-pass-but-not-sound = the
validity gap. A pre-registered codebook + full crossing (3 subject-teachers /
episode) is what raises inter-rater kappa from the LLM-judge's ~0.35 toward >=0.6.

Outputs: deliverables/packets_validity/packet_<subject>_T{1,2,3}.xlsx
"""
import hashlib
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from yandaojie_rounds import build_rounds, DELIVERABLES

TEACHERS = ["T1", "T2", "T3"]
SUBJECTS = ["数学", "科学", "英语"]
FONT = "Arial"
thin = Side(style="thin", color="C9C9C9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
FILL = {"id": "DDE6F2", "ctx": "EEF3FA", "label": "FFF6D5", "meta": "F0F0F0"}
HEAD = "1F3864"

MC_DV = '"已掌握,部分,未掌握"'
DEPTH_DV = '"扎实,对标签错机制,空洞或猜对,完全不懂,难判断"'
S15 = '"1,2,3,4,5"'

# (header, episode-key | None, width, dropdown-formula | None, fill)
COLS = [
    ("序号", "epid", 6, None, "id"),
    ("学生", "student", 6, None, "ctx"),
    ("课题", "topic", 13, None, "ctx"),
    ("本课学习目标", "objs", 30, None, "ctx"),
    ("学生反思（输入）", "refl", 22, None, "ctx"),
    ("完整辩护对话（AI追问 → 学生作答，逐轮）", "dialog", 62, None, "ctx"),
    ("①MC视角\n(只看答案/结论)", None, 13, MC_DV, "label"),
    ("②辩护深度\n(看完整辩护)", None, 15, DEPTH_DV, "label"),
    ("③辩护揭示了什么(引用一句)", None, 24, None, "label"),
    ("④信心(1–5)", None, 8, S15, "label"),
    ("评分老师", "__teacher__", 8, None, "meta"),
    ("备注", None, 12, None, "meta"),
]


# 8 norming episodes (positions in the (subject,sid)-sorted episode list) + facilitator gold
NORMING_POS = [5, 1, 0, 28, 14, 23, 32, 90]
GOLD = {
    5: ("已掌握", "扎实", "自我纠正后正确解释缩小(乘1/2不变形)与乘0边界，结论对+机制对。"),
    1: ("已掌握", "扎实", "能反向应用 (4,6)→(2,3) 乘1/2，说明真懂“同倍数”规则。临界：未口头解释为何形状不变，但能灵活应用 → 判扎实。"),
    0: ("已掌握", "对标签错机制", "说出“同时乘同一个数”(标签对)，但把不同倍数算成(4,9)却不识别这会变形 → 机制错。"),
    28: ("已掌握", "对标签错机制", "(2,8)→(4,16) 算对，却解释成“向上向右2格”=平移模型，不是缩放 → 典型对标签错机制。"),
    14: ("已掌握", "对标签错机制", "知道 (bx,by) 规则，但面对 A(1,1)→(2,2) vs B(1,3)→(2,5) 只反复说“第一个乘2第二个不是”，算不出 3→5 非×2、说不清形变。【边界教学点】他给了不完整/错的机制 → 对标签错机制；若**完全说不出任何机制**才归“空洞”。"),
    23: ("已掌握", "空洞或猜对", "说“横竖都放大两倍”(规则对)，但问 (3,4)×2 只重复原话、答不出 (6,8) → 给不出任何机制/无法应用 → 空洞。"),
    32: ("已掌握", "空洞或猜对", "水→水蒸气判物理、苹果变色判化学(标签对)，但每条理由都断在半句(“因为shui'bian'c”)，机制从未给出 → 空洞(亦可难判断)。"),
    90: ("部分", "难判断", "英语：“I did homework”“4 hores”“is so diffcult”是语言产出、非概念机制；②深度构念在语言科基本不适用 → 难判断。【边界教学点】这正是为何效度增益在英语≈0：没有可暴露的独立机制。"),
}
SOP_ROWS = [
    ("step", "1. 读手册", "先读「编码手册」sheet，重点记 ②辩护深度 的 5 档定义与 4 条判别规则。"),
    ("step", "2. 对齐练习(norming)", "三位老师各自**独立**标完「对齐练习」sheet 的 8 条 → 与主持人手中的 gold key 逐条对照 → 讨论分歧（重点：对标签错机制 vs 空洞、语言科→难判断）→ 统一理解。对齐后再开始正式标注。"),
    ("step", "3. 正式标注", "在「标注表」逐行**独立**标注 ①MC视角 ②辩护深度 ③揭示了什么(引用一句) ④信心。三位老师互不商量。"),
    ("rule", "原则", "①只看学生说出的答案；②看整段辩护；二者**独立**。只要追问下机制站不住，就**不算“扎实”**（哪怕答案全对）。信息不足 → 难判断 + 低信心。语言科若无可辩护的机制 → 难判断。"),
    ("rule", "节奏", "每条约 2–3 分钟；进入下一条后不回改上一条。"),
    ("rule", "完成", "交回文件；研究者计算评分员信度(κ)。若某维度 κ<0.6，回到 norming 再对齐一轮后重标该批。"),
]


def build_sop(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 78
    ws.cell(1, 1, "评分流程 SOP（标注前必读）").font = Font(name=FONT, bold=True, size=16, color="1F3864")
    r = 3
    for kind, a, b in SOP_ROWS:
        ws.cell(r, 1, a).font = Font(name=FONT, bold=True, size=12)
        ws.cell(r, 2, b).font = Font(name=FONT, size=11)
        ws.cell(r, 1).alignment = WRAP; ws.cell(r, 2).alignment = WRAP
        fg = "E7F4E4" if kind == "step" else "FFF6D5"
        ws.cell(r, 1).fill = PatternFill("solid", fgColor=fg); ws.cell(r, 2).fill = PatternFill("solid", fgColor=fg)
        r += 1


def build_episodes():
    rounds = build_rounds()
    by_sess = defaultdict(list)
    for r in rounds:
        by_sess[r["session_id"]].append(r)
    for s in by_sess:
        by_sess[s].sort(key=lambda r: r["round_index"] if r["round_index"] is not None else -1)
    eps = []
    for s, rs in by_sess.items():
        turns = [(r["round_index"], r["ai_question"], (r.get("student_answer") or "").strip()) for r in rs]
        turns = [t for t in turns if t[1]]
        if len([t for t in turns if t[2]]) >= 2:
            dialog = "\n".join(
                f"R{t[0]+1} AI：{t[1].strip()}\n    学生：「{t[2] or '（空）'}」" for t in turns)
            eps.append({
                "sid": s, "subject": rs[0]["subject"], "student": rs[0]["student_id"],
                "topic": rs[0]["topic"],
                "objs": "\n".join(f"{i+1}) {o}" for i, o in enumerate(rs[0]["objectives"])),
                "refl": rs[0]["reflection_text"], "dialog": dialog,
            })
    eps.sort(key=lambda e: (e["subject"], e["sid"]))
    for i, e in enumerate([x for x in eps]):
        pass
    return eps


def codebook(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 78
    r = 1
    def head(t):
        nonlocal r
        c = ws.cell(r, 1, t); c.font = Font(name=FONT, bold=True, size=13, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HEAD); ws.cell(r, 2).fill = PatternFill("solid", fgColor=HEAD)
        c.alignment = WRAP; r += 1
    def row(a, b, fill=None, bold=False):
        nonlocal r
        ws.cell(r, 1, a).font = Font(name=FONT, bold=True, size=11)
        ws.cell(r, 2, b).font = Font(name=FONT, bold=bold, size=11)
        ws.cell(r, 1).alignment = WRAP; ws.cell(r, 2).alignment = WRAP
        if fill:
            ws.cell(r, 1).fill = PatternFill("solid", fgColor=fill); ws.cell(r, 2).fill = PatternFill("solid", fgColor=fill)
        r += 1
    ws.cell(r, 1, "辩护深度标注 · 编码手册（Codebook）").font = Font(name=FONT, bold=True, size=16, color="1F3864"); r += 2
    head("你的任务")
    row("一行 = 一次辩护", "每行是一个学生就本课目标与 AI 的完整“答辩”对话（AI 追问 → 学生作答，逐轮）。读完整段，做两个独立判断：①MC视角 ②辩护深度。")
    row("为什么", "我们要量化：有多少“看答案像掌握了”的学生，其实一辩护就露馅。所以①只看答案、②看整段辩护，二者会故意分开。")
    r += 1
    head("① MC视角（只看学生说出的答案/结论，忽略论证质量）")
    row("已掌握", "学生最终给出的答案/结论是对的——若用选择题/对答案式评分会判对。", FILL["label"])
    row("部分", "答案部分对 / 不完整。", FILL["label"])
    row("未掌握", "答案错，或始终没给出答案。", FILL["label"])
    r += 1
    head("② 辩护深度（看完整个辩护过程）")
    row("扎实", "结论对，且能说出正确的机制/理由，追问下站得住。", "E7F4E4")
    row("对标签错机制", "结论/标签对，但一追问机制就暴露**错误或缺失的理解**——学生给出了一个（错的）解释。\n例：(2,8)→(4,16) 算对，却解释成“向上向右移2格”（平移模型，不是缩放）；判“化学变化”对，却说理由是“味道一样”。", "CDE7C8")
    row("空洞或猜对", "答案对/接近对，但**给不出任何机制**：循环重复、空、“不知道”、明显套规则或猜。\n例：会说“都放大两倍”，但问 (3,4)×2 答不出 (6,8)，只重复原话。", "CDE7C8")
    row("完全不懂", "通篇错。", "EEDDDD")
    row("难判断", "信息不足，无法归类（如全程乱码/空答）。", "F0F0F0")
    r += 1
    head("关键判别规则（请务必遵守 —— 这决定我们标注是否一致）")
    row("规则1", "区分“对标签错机制”vs“空洞或猜对”：前者学生**给出了一个错误的机制/解释**；后者学生**给不出任何机制**（重复/空/不知道）。")
    row("规则2", "只要追问下机制站不住，就**不算“扎实”**——哪怕答案全对。“扎实”要求结论对**且**理由对。")
    row("规则3", "①与②独立：一个学生完全可以 ①已掌握 + ②对标签错机制（这正是本研究的核心）。")
    row("规则4", "信心低/信息不足，②选“难判断”，别硬归类；④信心写 1–2。")
    r += 1
    head("对照例（先各自标，再核对）")
    row("例1（数学）", "学生算对 (2,8)→(4,16)，解释“会向上向右2格”。 → ①已掌握 ②对标签错机制")
    row("例2（数学）", "学生说“横纵都放大两倍”，但问 (3,4)×2 只会重复，答不出 (6,8)。 → ①已掌握 ②空洞或猜对")
    row("例3（科学）", "判苹果变色为“化学变化”（对），理由“变黄、味道一样”。 → ①已掌握 ②对标签错机制")
    row("例4（扎实对照）", "算对并解释“等比例放大、乘相同的数所以形状不变”。 → ①已掌握 ②扎实")
    r += 1
    row("正式标注前", "三位老师先用同样的 8 条样例做一次对齐（norming），统一对②各档的理解，再各自独立标注。", "FFF6D5", bold=True)


def write_sheet(ws, eps, teacher):
    ws.sheet_view.showGridLines = False
    for ci, (h, k, w, dv, fg) in enumerate(COLS, 1):
        c = ws.cell(1, ci, h); c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=HEAD); c.alignment = CTR; c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 40
    for ri, e in enumerate(eps, 2):
        for ci, (h, k, w, dv, fg) in enumerate(COLS, 1):
            c = ws.cell(ri, ci); c.border = BORDER; c.font = Font(name=FONT, size=10)
            if k == "__teacher__":
                c.value = teacher; c.fill = PatternFill("solid", fgColor=FILL["meta"]); c.alignment = CTR
            elif k:
                c.value = e.get(k, ""); c.fill = PatternFill("solid", fgColor=FILL["id"] if fg == "id" else FILL["ctx"])
                c.alignment = CTR if fg == "id" else WRAP
            else:
                c.fill = PatternFill("solid", fgColor=FILL[fg]); c.alignment = CTR if dv else WRAP
        ws.row_dimensions[ri].height = 150
    ws.freeze_panes = "B2"
    nrow = len(eps) + 1
    seen = {}
    for ci, (h, k, w, dv, fg) in enumerate(COLS, 1):
        if dv:
            d = DataValidation(type="list", formula1=dv, allow_blank=True, showErrorMessage=False)
            ws.add_data_validation(d); L = get_column_letter(ci); d.add(f"{L}2:{L}{nrow}")


def order_for(seed, eps):
    return sorted(eps, key=lambda e: hashlib.md5(f"{e['sid']}|{seed}".encode()).hexdigest())


def main():
    pdir = DELIVERABLES / "packets_validity"
    pdir.mkdir(parents=True, exist_ok=True)
    eps = build_episodes()
    # number within subject
    bysub = defaultdict(list)
    for e in eps:
        bysub[e["subject"]].append(e)
    for sub, xs in bysub.items():
        for i, e in enumerate(xs):
            e["epid"] = f"{sub[0]}{i+1}"
    norming = [eps[i] for i in NORMING_POS]
    for sub in SUBJECTS:
        sxs = bysub[sub]
        for t in TEACHERS:
            wb = Workbook()
            ws0 = wb.active; ws0.title = "编码手册"; codebook(ws0)
            build_sop(wb.create_sheet("评分SOP"))
            write_sheet(wb.create_sheet("对齐练习"), norming, t)   # 8 norming episodes, blank
            write_sheet(wb.create_sheet("标注表"), order_for(f"{sub}|{t}", sxs), t)
            wb.save(str(pdir / f"packet_{sub}_{t}.xlsx"))
        print(f"{sub}: {len(sxs)} episodes × 3 teachers")
    # facilitator gold key for the norming 8
    lines = ["# 对齐练习 (norming) · 主持人 Gold Key", "",
             "三位老师各自独立标完每份包里「对齐练习」sheet 的 8 条后，用下表对照、讨论分歧，达成一致再开始正式标注。",
             "", "| # | 序号 | 学科 | ①MC视角 | ②辩护深度 | 判定理由 / 边界教学点 |",
             "|---|---|---|---|---|---|"]
    for n, i in enumerate(NORMING_POS, 1):
        e = eps[i]; mc, dep, why = GOLD[i]
        lines.append(f"| {n} | {e['epid']} | {e['subject']} | {mc} | **{dep}** | {why} |")
    (DELIVERABLES / "norming_goldkey.md").write_text("\n".join(lines), encoding="utf-8")
    print(f"wrote {len(SUBJECTS)*len(TEACHERS)} packets (编码手册+评分SOP+对齐练习+标注表) + norming_goldkey.md")


if __name__ == "__main__":
    main()
