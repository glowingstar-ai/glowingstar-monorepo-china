"""Generate the v3 teacher annotation packets (4-label, two-screen).

Load-bearing reliability instrument for the reframed paper. Per defense episode
each subject-teacher supplies FOUR labels across TWO screens:

  Screen 1 (blind to the defense — sees ONLY the reflection + round-1 answer):
    L1  one-shot verdict          已掌握 / 部分 / 未掌握   (+ confidence)
        -> informationally-independent MC counterfactual (fixes the "MC is a
           reconstructed counterfactual" threat).

  Screen 2 (full defense transcript):
    L2  真实理解是否扎实           扎实 / 不扎实 / 难判断  (+ confidence)  <-- the kappa GATE
    L3  失效类别 (only if 不扎实)   对标签错机制 / 空洞或猜对 / 完全不懂   (3-way, confirmatory)
    L4  该重教什么 (单个点选)       针对性纠正 / 重讲整个概念 / 加强练习 / 补基础表达 / 无需  (= Study-2 H2 baseline)

All teacher-fill fields are dropdowns; every field is REQUIRED (③ is conditional on ②=不扎实).
The 6-type typology (T1..T6) is assigned by the AI/analysis layer, not the teacher.

Over-credit = L1==已掌握  AND  L2==不扎实.
Keyed by GLOBAL idx (0..105) so it joins to _validity_ratings.json (LLM-judge),
_typology_input.json (typology), and the future 3-AI scorer outputs.

Outputs: deliverables/packets_v3/packet_<subject>_T{1,2,3}.xlsx
         deliverables/norming_goldkey_v3.md
"""
import hashlib, json
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from yandaojie_rounds import DELIVERABLES

TEACHERS = ["T1", "T2", "T3"]
SUBJECTS = ["数学", "科学", "英语"]
FONT = "微软雅黑"  # CJK-safe on Chinese Windows; Excel substitutes gracefully elsewhere
thin = Side(style="thin", color="C9C9C9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
FILL = {"id": "DDE6F2", "ctx": "EEF3FA", "label": "FFF6D5", "meta": "F0F0F0"}
HEAD = "1F3864"

L1_DV = '"已掌握,部分,未掌握"'
L2_DV = '"扎实,不扎实,难判断"'
L3_DV = '"对标签错机制,空洞或猜对,完全不懂"'
# ④ 该重教什么 —— 单个点选，沿“泛泛重讲 vs 针对性纠正”轴（= Study 2 H2 的构念）
RETEACH_DV = '"针对性纠正错误机制,重讲整个概念,加强练习(执行),补基础/表达,无需/不适用"'
S15 = '"1,2,3,4,5"'

# ---- screen 1: only reflection + round-1 answer (blind to defense) ----
COLS_A = [
    ("序号", "idx", 7, None, "id"),
    ("学科", "subject", 6, None, "ctx"),
    ("课题", "topic", 12, None, "ctx"),
    ("本课学习目标", "objs", 26, None, "ctx"),
    ("学生反思（输入）", "refl", 20, None, "ctx"),
    ("第1轮 AI 提问", "r1q", 28, None, "ctx"),
    ("第1轮 学生作答", "r1a", 18, None, "ctx"),
    ("①只看此答案·对答案式评分会判", None, 14, L1_DV, "label"),
    ("①信心（1–5）", None, 8, S15, "label"),
    ("评分老师", "__teacher__", 8, None, "meta"),
]

# ---- screen 2: full defense transcript ----
COLS_B = [
    ("序号", "idx", 7, None, "id"),
    ("学科", "subject", 6, None, "ctx"),
    ("课题", "topic", 12, None, "ctx"),
    ("学生反思（输入）", "refl", 18, None, "ctx"),
    ("完整辩护对话（AI 追问 → 学生作答，逐轮）", "dialog", 58, None, "ctx"),
    ("②真实理解是否扎实", None, 12, L2_DV, "label"),
    ("③若不扎实·失效类别（3 类）", None, 14, L3_DV, "label"),
    ("④该重教什么（点选）", None, 20, RETEACH_DV, "label"),
    ("②信心（1–5）", None, 8, S15, "label"),
    ("评分老师", "__teacher__", 8, None, "meta"),
]

# 8 norming episodes (global idx) + facilitator gold in the v3 scheme
NORMING_IDX = [5, 1, 0, 28, 14, 23, 32, 90]
# idx: (L1①, L2②, L3③_3way, L3fine③细, L4④_reteach_MC, rationale/boundary note)
GOLD = {
    5:  ("已掌握", "扎实", "—", "—", "无需/不适用",
         "自我纠正后正确解释缩小(乘1/2不变形)与乘0边界，结论对+机制对 → 扎实。"),
    1:  ("已掌握", "扎实", "—", "—", "无需/不适用",
         "能反向应用 (4,6)→(2,3) 乘1/2，真懂“同倍数”。临界：未口头解释为何形状不变，但能灵活应用 → 仍判扎实。"),
    0:  ("已掌握", "不扎实", "对标签错机制", "T2错机制", "针对性纠正错误机制",
         "说“同时乘同一个数”(标签对)，但把不同倍数算成(4,9)且不识别会变形 → 给出错机制 → 对标签错机制 / T2。重教：等比例才保形。"),
    28: ("已掌握", "不扎实", "对标签错机制", "T2错机制", "针对性纠正错误机制",
         "(2,8)→(4,16) 算对，却解释成“向上向右2格”=平移模型 → 典型对标签错机制 / T2。重教：缩放≠平移。"),
    14: ("已掌握", "不扎实", "对标签错机制", "T1背规则", "针对性纠正错误机制",
         "知道 (bx,by) 规则，但面对 A(1,1)→(2,2) 与 B(1,3)→(2,5) 只反复说“第一个乘2第二个不是”，给不出完整机制。"
         "【边界】给了不完整机制 → 对标签错机制；若**完全说不出任何机制**才归空洞。细分 T1/T2 皆可，记 T1。"),
    23: ("已掌握", "不扎实", "空洞或猜对", "T6不会执行", "加强练习(执行)",
         "说“横竖都放大两倍”(规则对)，但问 (3,4)×2 只重复原话、答不出 (6,8) → 给不出可执行机制 → 空洞；细分 T6(会说规则不会执行)。"),
    32: ("已掌握", "不扎实", "空洞或猜对", "T4碎片", "补基础/表达",
         "水→水蒸气判物理、苹果变色判化学(标签对)，但理由都断在半句(“因为shui'bian'c”)，机制从未给出 → 空洞；细分 T4(碎片停滞)。可选难判断。"),
    90: ("部分", "难判断", "—", "—", "无需/不适用",
         "“I did homework”“4 hores”“is so diffcult”是语言产出、非概念机制 → 难判断。"
         "【边界】这正是为何效度增益在英语≈0：没有可暴露的独立机制。"),
}

SOP_ROWS = [
    ("step", "1. 读手册", "先读「编码手册」这一页：①②③④ 四项标注的定义与判别规则；重点记 ②「扎实」的门槛与 ③ 三类的区分。"),
    ("step", "2. 对齐练习", "三位老师各自**独立**标完「对齐练习」8 条 → 与主持人手中的标准答案逐条对照 → 讨论分歧（重点：对标签错机制 与 空洞 的区分；语言科 → 难判断）→ 统一理解，再开始正式标注。"),
    ("step", "3. 先做屏1（只看答案）", "打开「屏1·只看答案」：**只看学生反思 + 第1轮作答**，据此填 ①（对答案式/选择题评分会判哪类）。**做完全部屏1 再打开屏2**——屏1 是模拟“没有后续追问、只看答案”的情形，不要参考屏2。"),
    ("step", "4. 再做屏2（看完整辩护）", "打开「屏2·看完整答辩」：读整段辩护，**点选** ② 是否扎实、③ 失效类别（仅当 ② 选“不扎实”时填；②选扎实/难判断则 ③ 留空）、④ 该重教什么。全部是下拉框选项，不用打字。"),
    ("rule", "①②独立", "① 只看答案、② 看整段辩护，二者**独立**。一个学生完全可以 ① 已掌握 + ② 不扎实——这正是本研究要量化的“效度缺口”。"),
    ("rule", "“扎实”门槛", "只要追问下机制站不住，就**不算扎实**（哪怕答案全对）。扎实 = 结论对**且**机制对。信息不足 → ② 难判断 + 低信心。"),
    ("rule", "语言科", "英语若无可辩护的独立概念机制（只是语言产出）→ ② 难判断，④ 选“无需/不适用”。"),
    ("rule", "节奏 / 独立", "每条约 2–4 分钟；进入下一条不回改上一条；三位老师互不商量。可分次完成，但同一页一次做完更一致。"),
    ("rule", "完成", "交回文件；研究者计算三位老师的一致性。若 ②（扎实／不扎实）或 ③ 的一致性不达标，回到对齐练习再对齐一轮后重标该批。"),
]


def load_episodes():
    eps = json.load(open(DELIVERABLES / "_validity_episodes.json"))
    out = {}
    for e in eps:
        turns = e.get("turns") or []
        t1 = turns[0] if turns else {}
        dialog = "\n".join(
            f"R{t.get('round')} AI：{(t.get('ai_q') or '').strip()}\n    学生：「{(t.get('student') or '').strip() or '（空）'}」"
            for t in turns if (t.get('ai_q'))
        )
        out[e["idx"]] = {
            "idx": e["idx"], "subject": e["subject"], "topic": e.get("topic", ""),
            "objs": "\n".join(f"{i+1}) {o}" for i, o in enumerate(e.get("objectives") or [])),
            "refl": e.get("reflection", ""),
            "r1q": (t1.get("ai_q") or "").strip(),
            "r1a": (t1.get("student") or "").strip() or "（空）",
            "dialog": dialog,
        }
    return out


def _blockhead(ws, r, t):
    c = ws.cell(r, 1, t); c.font = Font(name=FONT, bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=HEAD); ws.cell(r, 2).fill = PatternFill("solid", fgColor=HEAD)
    c.alignment = WRAP


def codebook(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 84
    r = 1

    def head(t):
        nonlocal r
        _blockhead(ws, r, t); r += 1

    def row(a, b, fill=None, bold=False):
        nonlocal r
        ws.cell(r, 1, a).font = Font(name=FONT, bold=True, size=11)
        ws.cell(r, 2, b).font = Font(name=FONT, bold=bold, size=11)
        ws.cell(r, 1).alignment = WRAP; ws.cell(r, 2).alignment = WRAP
        if fill:
            ws.cell(r, 1).fill = PatternFill("solid", fgColor=fill); ws.cell(r, 2).fill = PatternFill("solid", fgColor=fill)
        r += 1

    ws.cell(r, 1, "辩护深度标注 · 编码手册").font = Font(name=FONT, bold=True, size=16, color="1F3864"); r += 2
    head("你的任务（一行 = 一次辩护）")
    row("两屏、四项标注", "屏1 只看反思 + 第1轮答案 → 填 ①；屏2 看整段辩护 → 填 ②③④。目的：量化“看答案像掌握了、一追问就露馅”的比例。")

    head("① 一次性/对答案式评分（屏1：只看第1轮答案，先不看后续追问）")
    row("已掌握", "只看这一句答案/结论，若用选择题/对答案式评分会判对。", FILL["label"])
    row("部分", "答案部分对/不完整。", FILL["label"])
    row("未掌握", "答案错，或没给出答案。", FILL["label"])

    head("② 真实理解是否扎实（屏2：看整段辩护）—— 最关键的一项，务必标准一致")
    row("扎实", "结论对，**且**能说出正确机制/理由，追问下站得住。", "E7F4E4")
    row("不扎实", "结论/答案看着对或接近对，但追问下机制暴露问题（错误、缺失、或自相矛盾）。**只要机制站不住就选这个，哪怕答案全对。**", "CDE7C8")
    row("难判断", "信息不足无法归类（全程乱码/空答；或语言科无独立概念机制可辩护）。", "F0F0F0")

    head("③ 失效类别（仅当 ② 选“不扎实”；3 类，必填）")
    row("对标签错机制", "标签/结论对，但学生**给出了一个错误的机制/解释**。例：判“化学变化”对，理由是“味道一样”；算对 (2,8)→(4,16) 却说是“平移2格”。", "CDE7C8")
    row("空洞或猜对", "答案对/接近对，但**给不出任何机制**：循环重复、空、“不知道”、套规则却不会用。", "CDE7C8")
    row("完全不懂", "通篇错、或无任何可评估内容。", "EEDDDD")
    row("★区分要点", "“对标签错机制”= 给了**错的**机制；“空洞”= **给不出**机制。这条最影响一致性。", FILL["label"], bold=True)

    head("④ 该重教什么（屏2；点选，必填）—— 全部是选项，不用写字")
    row("针对性纠正错误机制", "点出他具体哪个机制错了/缺了，只补那一处。（多用于 ③ 对标签错机制）", "E7F4E4")
    row("重讲整个概念", "机制整体没建立，需要把这个概念从头再讲一遍。", FILL["label"])
    row("加强练习(执行)", "会说规则/概念，但做不出、组合不出——需要多练。", FILL["label"])
    row("补基础/表达", "卡在基础或语言表达，给不出完整命题。", FILL["label"])
    row("无需/不适用", "② 已判扎实无需处理；或语言科无独立机制可补。", FILL["label"])

    head("关键规则（决定标注一致性）")
    row("规则1", "① 与 ② 独立：可以 ① 已掌握 + ② 不扎实。")
    row("规则2", "“扎实”要求结论对**且**机制对；机制站不住即不扎实。")
    row("规则3", "③“对标签错机制”是给了错机制；“空洞”是给不出机制——这条最影响一致性。")
    row("规则4", "信息不足/语言科无机制 → ② 难判断，别硬归类，信心写 1–2。")
    row("正式标注前", "三位老师先用同样的 8 条做对齐练习，统一理解再各自独立标。", FILL["label"], bold=True)


def build_sop(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 86
    ws.cell(1, 1, "评分流程（标注前必读）").font = Font(name=FONT, bold=True, size=16, color="1F3864")
    r = 3
    for kind, a, b in SOP_ROWS:
        ws.cell(r, 1, a).font = Font(name=FONT, bold=True, size=12)
        ws.cell(r, 2, b).font = Font(name=FONT, size=11)
        ws.cell(r, 1).alignment = WRAP; ws.cell(r, 2).alignment = WRAP
        fg = "E7F4E4" if kind == "step" else "FFF6D5"
        ws.cell(r, 1).fill = PatternFill("solid", fgColor=fg); ws.cell(r, 2).fill = PatternFill("solid", fgColor=fg)
        r += 1


def write_sheet(ws, cols, eps, teacher, row_h=150):
    ws.sheet_view.showGridLines = False
    for ci, (h, k, w, dv, fg) in enumerate(cols, 1):
        c = ws.cell(1, ci, h); c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=HEAD); c.alignment = CTR; c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 42
    for ri, e in enumerate(eps, 2):
        for ci, (h, k, w, dv, fg) in enumerate(cols, 1):
            c = ws.cell(ri, ci); c.border = BORDER; c.font = Font(name=FONT, size=10)
            if k == "__teacher__":
                c.value = teacher; c.fill = PatternFill("solid", fgColor=FILL["meta"]); c.alignment = CTR
            elif k:
                c.value = e.get(k, ""); c.fill = PatternFill("solid", fgColor=FILL["id"] if fg == "id" else FILL["ctx"])
                c.alignment = CTR if fg == "id" else WRAP
            else:
                c.fill = PatternFill("solid", fgColor=FILL[fg]); c.alignment = CTR if dv else WRAP
        ws.row_dimensions[ri].height = row_h
    ws.freeze_panes = "B2"
    nrow = len(eps) + 1
    for ci, (h, k, w, dv, fg) in enumerate(cols, 1):
        if dv:
            d = DataValidation(type="list", formula1=dv, allow_blank=True, showErrorMessage=False)
            ws.add_data_validation(d); L = get_column_letter(ci); d.add(f"{L}2:{L}{nrow}")


def order_for(seed, eps):
    return sorted(eps, key=lambda e: hashlib.md5(f"{e['idx']}|{seed}".encode()).hexdigest())


def main():
    pdir = DELIVERABLES / "packets_v3"
    pdir.mkdir(parents=True, exist_ok=True)
    EP = load_episodes()
    bysub = defaultdict(list)
    for idx, e in EP.items():
        bysub[e["subject"]].append(e)
    norming = [EP[i] for i in NORMING_IDX]

    for sub in SUBJECTS:
        sxs = bysub[sub]
        for t in TEACHERS:
            wb = Workbook()
            ws0 = wb.active; ws0.title = "编码手册"; codebook(ws0)
            build_sop(wb.create_sheet("评分流程"))
            # norming: combined practice on the full transcript (both screens' labels)
            write_sheet(wb.create_sheet("对齐练习"), COLS_B, norming, t)
            order = order_for(f"{sub}|{t}", sxs)
            write_sheet(wb.create_sheet("屏1·只看答案"), COLS_A, order, t, row_h=110)
            write_sheet(wb.create_sheet("屏2·看完整答辩"), COLS_B, order, t)
            wb.save(str(pdir / f"packet_{sub}_{t}.xlsx"))
        print(f"{sub}: {len(sxs)} episodes × 3 teachers")

    # facilitator gold key (v3 scheme)
    lines = ["# 对齐练习 · 主持人标准答案（四项标注）", "",
             "三位老师各自独立标完「对齐练习」8 条后，用下表逐条对照、讨论分歧，达成一致再开始正式标注。",
             "「序号」为全局编号，与题库及后续打分对齐。", "",
             "| # | 序号 | 学科 | ①只看答案 | ②是否扎实 | ③三类 | ④该重教什么 | 判定理由·边界 |",
             "|---|---|---|---|---|---|---|---|"]
    for n, i in enumerate(NORMING_IDX, 1):
        e = EP[i]; L1, L2, L3, _L3f, rt, why = GOLD[i]
        lines.append(f"| {n} | {i} | {e['subject']} | {L1} | **{L2}** | {L3} | {rt} | {why} |")
    (DELIVERABLES / "norming_goldkey_v3.md").write_text("\n".join(lines), encoding="utf-8")
    print(f"wrote {len(SUBJECTS)*len(TEACHERS)} packets to {pdir.name}/ "
          f"(编码手册+评分流程+对齐练习+屏1+屏2) + norming_goldkey_v3.md")


if __name__ == "__main__":
    main()
