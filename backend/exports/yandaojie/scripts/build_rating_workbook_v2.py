"""Build the THREE-PHASE teacher rating workbook (v2) from the real-only export.

Phase 1  盲评           — teacher's own diagnosis + question + confidence (no AI)
Phase 2  看AI后重评      — reveal AI diagnosis (±reasoning), teacher RE-RATES own
                          diagnosis + confidence  -> measures AI influence (WOA)
Phase 3  评AI反馈质量    — rate the AI's diagnosis / question / + student outcome

The ±reasoning manipulation is realised in-sheet: the "AI推理过程" column is
populated ONLY for rounds whose 条件 == "+推理"; "仅结论" rounds see diagnosis only.
"""
import json
from collections import defaultdict
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from yandaojie_rounds import RAW, DELIVERABLES  # robust base resolution

OUT = str(DELIVERABLES / "teacher_rating_workbook_v2.xlsx")


def load(n):
    return [json.loads(l) for l in open(RAW / f"{n}.jsonl") if l.strip()]


sessions = load("sessions"); questions = load("defense_questions"); turns = load("defense_turns")
sess = {s["session_id"]: s for s in sessions}
ans = {(t.get("session_id"), t.get("round_index")): t for t in turns}


def reflen(s):
    return sum(len((r.get("learned") or "").strip()) for r in (s.get("reflections") or []))


_lens = sorted(reflen(s) for s in sessions)
TB1, TB2 = _lens[len(_lens) // 3], _lens[2 * len(_lens) // 3]
def band(s):
    n = reflen(s); return "薄" if n <= TB1 else ("中" if n <= TB2 else "厚")


def fmt_obj(o): return "\n".join(f"{i+1}) {x}" for i, x in enumerate(o)) if o else ""
def fmt_refl(s):
    out = []
    for r in (s.get("reflections") or []):
        idx = r.get("objective_index"); tag = f"目标{idx+1}" if isinstance(idx, int) else "目标?"
        learned = (r.get("learned") or "").strip() or "（空）"; q = (r.get("questions") or "").strip()
        out.append(f"{tag}：学到=「{learned}」" + (f"；疑问=「{q}」" if q else ""))
    return "\n".join(out)
def fmt_diag(d, key):
    items = (d or {}).get(key) or []; p = "M" if key == "mastered" else "N"
    return "\n".join(f"{p}{i+1}. {x}" for i, x in enumerate(items))
def fmt_tgt(q):
    objs = sess.get(q.get("session_id"), {}).get("learning_objectives") or []; out = []
    for t in (q.get("targeted_objectives") or []):
        idx = t.get("objective_index")
        out.append(f"目标{(idx+1) if isinstance(idx,int) else '?'}：{t.get('reason') or ''}")
    return "\n".join(out)


by_sess = defaultdict(list)
for q in questions:
    by_sess[q.get("session_id")].append(q)
for sid in by_sess:
    by_sess[sid].sort(key=lambda q: (q.get("round_index") if q.get("round_index") is not None else -1))

rows = []
for sid, qs in by_sess.items():
    s = sess.get(sid, {}); objs = s.get("learning_objectives") or []; prev = None
    for q in qs:
        ri = q.get("round_index"); a = ans.get((sid, ri)); at = (a or {}).get("answer_text")
        rows.append({
            "round_id": f"{sid[:6]}-R{(ri+1) if isinstance(ri,int) else '?'}",
            "student_id": s.get("student_id"), "subject": s.get("subject_label"),
            "topic": s.get("subject_topic"), "round": (ri + 1) if isinstance(ri, int) else "",
            "objectives": fmt_obj(objs), "reflection": fmt_refl(s),
            "prev_q": (prev or {}).get("q", ""), "prev_a": (prev or {}).get("a", ""),
            "ai_mastered": fmt_diag(q.get("diagnoses"), "mastered"),
            "ai_not_mastered": fmt_diag(q.get("diagnoses"), "not_mastered"),
            "ai_targets": fmt_tgt(q), "ai_question": q.get("question") or "",
            "student_answer": at or "", "reasoning": q.get("reasoning_content") or "",
            "band": band(s),
            "stratum": f"{s.get('subject_label')}×R{(ri+1) if isinstance(ri,int) else '?'}×{band(s)}",
        })
        prev = {"q": q.get("question") or "", "a": at or ""}

# stratified ~140 sample (oversample thin)
strata = defaultdict(list)
for i, r in enumerate(rows):
    strata[r["stratum"]].append(i)
by_band = defaultdict(list)
for i, r in enumerate(rows):
    by_band[r["band"]].append(i)
for b in by_band:
    by_band[b].sort(key=lambda i: (rows[i]["stratum"], i))
TARGET = 140; QUOTA = {"薄": 0.50, "中": 0.30, "厚": 0.20}
sample_idx = set()
for b, idxs in by_band.items():
    quota = min(len(idxs), max(1, round(TARGET * QUOTA.get(b, 0.2)))); step = len(idxs) / quota
    for k in range(quota):
        sample_idx.add(idxs[int(k * step)])
for i, r in enumerate(rows):
    r["sample"] = "是" if i in sample_idx else "否"

# ±reasoning condition: balanced within each stratum (alternate)
for st, idxs in strata.items():
    for pos, i in enumerate(idxs):
        rows[i]["cond"] = "+推理" if pos % 2 == 0 else "仅结论"

# ---------------- styling ----------------
FONT = "Arial"
FILLS = {
    "id": PatternFill("solid", fgColor="DDE6F2"),
    "ctx": PatternFill("solid", fgColor="EEF3FA"),
    "p1": PatternFill("solid", fgColor="E7F4E4"),   # phase 1 blind (green)
    "p2": PatternFill("solid", fgColor="CDE7C8"),   # phase 2 revision (deeper green)
    "q": PatternFill("solid", fgColor="FFF6D5"),    # phase 3 AI quality (yellow)
    "meta": PatternFill("solid", fgColor="F0F0F0"),
}
HEAD = PatternFill("solid", fgColor="1F3864")
P1HEAD = PatternFill("solid", fgColor="2E7D32")
P2HEAD = PatternFill("solid", fgColor="1B5E20")
QHEAD = PatternFill("solid", fgColor="B8860B")
thin = Side(style="thin", color="C9C9C9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
DROPDOWN_KINDS = {"S15", "DIAG_GND", "CONSIST", "ELICIT", "YESNO", "ATTITUDE", "ICAP", "CORRECT", "TEACHER", "FAB"}

wb = Workbook()

# ============ Sheet 1: 使用说明 ============
ws0 = wb.active; ws0.title = "使用说明"; ws0.sheet_view.showGridLines = False
ws0.column_dimensions["A"].width = 24; ws0.column_dimensions["B"].width = 26; ws0.column_dimensions["C"].width = 74
def w0(r, a, b="", c="", bold=False, head=False, fill=None):
    ws0.cell(r, 1, a); ws0.cell(r, 2, b); ws0.cell(r, 3, c)
    for col in (1, 2, 3):
        cc = ws0.cell(r, col)
        cc.font = Font(name=FONT, bold=bold or head, size=14 if head else 11, color="FFFFFF" if head else "000000")
        cc.alignment = WRAP
        if head: cc.fill = HEAD
        elif fill: cc.fill = fill
r = 1
ws0.cell(r, 1, "研道街 · 三阶段教师评分工作表（v2）"); ws0.cell(r, 1).font = Font(name=FONT, bold=True, size=16, color="1F3864"); r += 2
w0(r, "三个阶段（务必按顺序）", head=True); r += 1
for ln in [
    "阶段1 盲评（绿色）：在看不到 AI 的情况下，给出您自己的逐目标诊断、您会怎么追问、以及您的信心(1–5)。",
    "阶段2 看 AI 后重评（深绿）：揭示 AI 的诊断〔部分行附带 AI 推理过程〕。读完后，请【重新】给出您自己的诊断和信心——可以和阶段1一样，也可以改。我们要测 AI 把您的判断移动了多少。",
    "阶段3 评 AI 反馈质量（黄色）：再揭示 AI 追问与学生回答，对 AI 的诊断/追问/学生表现打分。",
    "最后：到「推理审计」表评 AI 推理本身（逻辑/faithfulness/有无捏造）。",
]:
    w0(r, "", "", ln); r += 1
r += 1
w0(r, "⚠ 关于「±推理」条件", head=True); r += 1
w0(r, "", "", "「条件」列标明本行是「+推理」还是「仅结论」。+推理 行会在「AI推理过程」列显示 AI 的推理；仅结论 行该列留空——这是实验操纵，请勿去别处找推理来看。两种条件的分配已配平。"); r += 1
w0(r, "", "", "阶段2 重评请基于您【当时屏幕上看到的】内容：+推理行可参考推理，仅结论行只看 AI 诊断结论。"); r += 2
w0(r, "⚠ 关于盲评", head=True); r += 1
w0(r, "", "", "Excel 无法强制隐藏。请自觉先填完阶段1绿色列，再向右；严格盲评建议改用问卷工具（可另配）。"); r += 2

w0(r, "评分图例（各列含义）", head=True); r += 1
ws0.cell(r, 1, "列"); ws0.cell(r, 2, "取值"); ws0.cell(r, 3, "含义")
for col in (1, 2, 3):
    ws0.cell(r, col).font = Font(name=FONT, bold=True, color="FFFFFF"); ws0.cell(r, col).fill = PatternFill("solid", fgColor="2E5496"); ws0.cell(r, col).alignment = WRAP
r += 1
legend = [
    ("【P1】盲诊_掌握度", "3/2/1/0", "3已掌握 2部分 1未掌握 0无法判断。逐目标写，如“目标1=2; 目标2=0”。", "p1"),
    ("【P1】我会怎么问", "文本", "如果由您来追问，您会问什么。", "p1"),
    ("【P1】信心", "1–5", "您对自己上面这个诊断有多确信。5=非常确信。", "p1"),
    ("【P2】重评_掌握度", "3/2/1/0", "读完 AI 后，您【现在】对每个目标的判断（可同可改）。", "p2"),
    ("【P2】信心", "1–5", "读完 AI 后您的信心。", "p2"),
    ("【P2】改判说明", "文本(可选)", "如果改了，为什么改；尽量写您依据的是 AI 的哪句话。", "p2"),
    ("【Q】诊断groundedness", "全部有据/个别臆造/多数臆造", "AI 每条掌握判断是否有反思原文支持。", "q"),
    ("【Q】与盲诊(P1)一致性", "一致/部分一致/矛盾", "AI 诊断与您【阶段1盲评】的异同。", "q"),
    ("【Q】诊断有用性", "1–5", "5精准可教 3偏泛 1没用/误导。", "q"),
    ("【Q】追问相关性", "1–5", "5紧扣本轮反思 1跑题/引用学生没说过的。", "q"),
    ("【Q】对准薄弱点", "1–5", "5正好戳中薄弱处 1针对已会的/不存在的问题。", "q"),
    ("【Q】引出力", "L1/L2/L3/L4", "L1回忆 L2解释 L3论证/辩护 L4迁移。L3/L4才算“答辩式”。", "q"),
    ("【Q】适切", "1–5", "对该学段难度是否合适。", "q"),
    ("【Q】会用吗", "是/否", "您会不会真对这孩子问这句。", "q"),
    ("【Q】作答态度", "真诚尝试/敷衍/空答或乱码", "学生是否认真作答。", "q"),
    ("【Q】ICAP", "被动/主动/建构/互动", "被动复述；主动操作无解释；建构自己的解释；互动在追问上修正深化。", "q"),
    ("【Q】深度", "1–5", "评有没有展开推理，不是对错。", "q"),
    ("【Q】正确性", "正确/部分正确/错误", "背景项，可选。", "q"),
    ("【Q】整轮推进", "1–5", "这一轮有没有逼出之前看不到的思考。", "q"),
    ("【Q】匹配度", "1–5", "AI 追问与学生水平匹配度。", "q"),
    ("（推理审计表）6a/6b/6c", "1–5 / 有无", "AI 推理逻辑、是否真支撑结论(faithfulness)、有无捏造。", "p2"),
]
for name, vals, desc, fg in legend:
    w0(r, name, vals, desc, fill=FILLS[fg]); ws0.cell(r, 1).font = Font(name=FONT, bold=True, size=11); r += 1

# ============ Sheet 2: 评分表 ============
ws = wb.create_sheet("评分表"); ws.sheet_view.showGridLines = False
# (header, key, width, dvkind, fill)
COLS = [
    ("round_id", "round_id", 11, None, "id"),
    ("学生ID", "student_id", 7, None, "ctx"),
    ("学科", "subject", 7, None, "ctx"),
    ("课题", "topic", 15, None, "ctx"),
    ("轮次", "round", 5, None, "ctx"),
    ("本课学习目标", "objectives", 34, None, "ctx"),
    ("学生反思（输入）", "reflection", 30, None, "ctx"),
    ("上一轮·AI问题", "prev_q", 20, None, "ctx"),
    ("上一轮·学生答", "prev_a", 16, None, "ctx"),
    # ---- Phase 1 ----
    ("【P1】盲诊_掌握度(3/2/1/0)", None, 16, None, "p1"),
    ("【P1】我会怎么问", None, 20, None, "p1"),
    ("【P1】信心", None, 7, "S15", "p1"),
    # ---- AI diagnosis reveal (start of Phase 2) ----
    ("AI诊断·已掌握", "ai_mastered", 26, None, "ctx"),
    ("AI诊断·未掌握", "ai_not_mastered", 26, None, "ctx"),
    ("AI针对目标(理由)", "ai_targets", 24, None, "ctx"),
    ("AI推理过程〔仅+推理条件〕", "__reasoning_cond__", 40, None, "ctx"),
    # ---- Phase 2 revision ----
    ("【P2】重评_掌握度(3/2/1/0)", None, 16, None, "p2"),
    ("【P2】信心", None, 7, "S15", "p2"),
    ("【P2】改判说明(可选)", None, 20, None, "p2"),
    # ---- Phase 3: AI quality ----
    ("【Q】诊断groundedness", None, 14, "DIAG_GND", "q"),
    ("【Q】臆造说明", None, 16, None, "q"),
    ("【Q】与盲诊(P1)一致性", None, 12, "CONSIST", "q"),
    ("【Q】诊断有用性", None, 8, "S15", "q"),
    ("AI追问", "ai_question", 30, None, "ctx"),
    ("【Q】追问相关性", None, 8, "S15", "q"),
    ("【Q】对准薄弱点", None, 9, "S15", "q"),
    ("【Q】引出力", None, 9, "ELICIT", "q"),
    ("【Q】适切", None, 7, "S15", "q"),
    ("【Q】会用吗", None, 8, "YESNO", "q"),
    ("【Q】会用原因", None, 14, None, "q"),
    ("学生回答", "student_answer", 24, None, "ctx"),
    ("【Q】作答态度", None, 12, "ATTITUDE", "q"),
    ("【Q】ICAP", None, 10, "ICAP", "q"),
    ("【Q】深度", None, 7, "S15", "q"),
    ("【Q】正确性", None, 11, "CORRECT", "q"),
    ("【Q】整轮推进", None, 9, "S15", "q"),
    ("【Q】匹配度", None, 8, "S15", "q"),
    # ---- meta ----
    ("评分老师", None, 9, "TEACHER", "meta"),
    ("抽样", "sample", 6, None, "meta"),
    ("条件", "cond", 9, None, "meta"),
    ("分层", "stratum", 15, None, "meta"),
    ("备注", None, 16, None, "meta"),
]
# header
for ci, (hdr, key, width, dv, fg) in enumerate(COLS, start=1):
    c = ws.cell(1, ci, hdr)
    c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    c.fill = P1HEAD if fg == "p1" else P2HEAD if fg == "p2" else QHEAD if fg == "q" else HEAD
    c.alignment = CTR; c.border = BORDER
ws.row_dimensions[1].height = 46
# data
for ri, row in enumerate(rows, start=2):
    for ci, (hdr, key, width, dv, fg) in enumerate(COLS, start=1):
        c = ws.cell(ri, ci); c.border = BORDER; c.font = Font(name=FONT, size=10)
        if key == "__reasoning_cond__":
            c.value = row["reasoning"] if row["cond"] == "+推理" else "〔仅结论条件：本行不提供推理〕"
            c.fill = FILLS["ctx"]; c.alignment = WRAP
        elif key is not None:
            c.value = row.get(key, ""); c.fill = FILLS[fg]
            if fg == "id": c.font = Font(name=FONT, size=10, bold=True); c.alignment = CTR
            else: c.alignment = CTR if fg == "meta" and key in ("sample", "cond") else WRAP
        else:
            c.fill = FILLS[fg]
            c.alignment = CTR if dv in DROPDOWN_KINDS else WRAP
for ci, (hdr, key, width, dv, fg) in enumerate(COLS, start=1):
    ws.column_dimensions[get_column_letter(ci)].width = width
ws.freeze_panes = "B2"
for ri in range(2, len(rows) + 2):
    ws.row_dimensions[ri].height = 96
# dropdowns
nrow = len(rows) + 1
DV = {"S15": '"1,2,3,4,5"', "DIAG_GND": '"全部有据,个别臆造,多数臆造"', "CONSIST": '"一致,部分一致,矛盾"',
      "ELICIT": '"L1回忆,L2解释,L3论证,L4迁移"', "YESNO": '"是,否"', "ATTITUDE": '"真诚尝试,敷衍,空答或乱码"',
      "ICAP": '"被动,主动,建构,互动"', "CORRECT": '"正确,部分正确,错误"', "TEACHER": '"T1,T2,T3"'}
for kind, formula in DV.items():
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv)
    for ci, (hdr, key, width, k, fg) in enumerate(COLS, start=1):
        if k == kind:
            L = get_column_letter(ci); dv.add(f"{L}2:{L}{nrow}")

# ============ Sheet 3: 推理审计 ============
ws2 = wb.create_sheet("推理审计"); ws2.sheet_view.showGridLines = False
RCOLS = [
    ("round_id", "round_id", 11, None, "id"), ("学科", "subject", 7, None, "ctx"),
    ("课题", "topic", 14, None, "ctx"), ("轮次", "round", 5, None, "ctx"),
    ("学生反思（输入）", "reflection", 28, None, "ctx"),
    ("AI诊断·已掌握", "ai_mastered", 24, None, "ctx"), ("AI诊断·未掌握", "ai_not_mastered", 24, None, "ctx"),
    ("AI追问", "ai_question", 26, None, "ctx"), ("AI推理过程（reasoning）", "reasoning", 52, None, "ctx"),
    ("⑥a推理逻辑", None, 9, "S15", "p2"), ("⑥b faithfulness", None, 11, "S15", "p2"),
    ("⑥c有无捏造", None, 10, "FAB", "p2"), ("评分老师", None, 9, "TEACHER", "meta"),
    ("条件", "cond", 9, None, "meta"), ("抽样", "sample", 6, None, "meta"), ("备注", None, 16, None, "meta"),
]
for ci, (hdr, *_x) in enumerate(RCOLS, start=1):
    c = ws2.cell(1, ci, hdr); c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    c.fill = P2HEAD if _x[-1] == "p2" else HEAD; c.alignment = CTR; c.border = BORDER
ws2.row_dimensions[1].height = 40
for ri, row in enumerate(rows, start=2):
    for ci, (hdr, key, width, dv, fg) in enumerate(RCOLS, start=1):
        c = ws2.cell(ri, ci); c.border = BORDER; c.font = Font(name=FONT, size=10)
        if key is not None:
            c.value = row.get(key, ""); c.fill = FILLS["id"] if fg == "id" else FILLS["ctx"] if fg == "ctx" else FILLS["meta"]
            c.alignment = CTR if (fg == "id" or key in ("sample", "cond")) else WRAP
        else:
            c.fill = FILLS[fg]; c.alignment = CTR
for ci, (hdr, key, width, dv, fg) in enumerate(RCOLS, start=1):
    ws2.column_dimensions[get_column_letter(ci)].width = width
ws2.freeze_panes = "B2"
for ri in range(2, len(rows) + 2):
    ws2.row_dimensions[ri].height = 110
nrow2 = len(rows) + 1
for kind, formula in {"S15": '"1,2,3,4,5"', "FAB": '"有,无"', "TEACHER": '"T1,T2,T3"'}.items():
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showErrorMessage=False)
    ws2.add_data_validation(dv)
    for ci, (hdr, key, width, k, fg) in enumerate(RCOLS, start=1):
        if k == kind:
            L = get_column_letter(ci); dv.add(f"{L}2:{L}{nrow2}")

wb.save(OUT)
ns = sum(1 for r in rows if r["sample"] == "是")
nc = sum(1 for r in rows if r["cond"] == "+推理")
print(f"wrote {OUT}: {len(rows)} rounds | sample={ns} | +推理={nc} 仅结论={len(rows)-nc} | cols={len(COLS)}")
