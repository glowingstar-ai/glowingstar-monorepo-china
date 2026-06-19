"""Analysis skeleton for the Yandaojie Design-C staged-disclosure expert audit.

Reads the FILLED per-subject packets (deliverables/packets/packet_<subject>_T*.xlsx)
and computes the staged influence metrics. Runs end-to-end on EMPTY packets too:
it reports how many ratings are present and skips metrics with no data yet.

Stages per teacher-round:  P1 盲评 -> P2a 看AI结论后 -> P2b 再评 -> P3 评AI质量
  推理臂 (reasoning arm): P2b after seeing AI reasoning
  控制臂 (control arm):   P2b re-rate with no new info  (re-exposure baseline)

Metrics:
  1. Advice effect (P1->P2a): how much the bare AI recommendation moves experts.
  2. Staged increment (P2a->P2b) by arm + NET reasoning effect = 推理臂 − 控制臂
     (diff-in-diff: isolates the reasoning trace from mere re-rating).
  3. Confidence: ΔConf at each step; net reasoning ΔConf.
  4. Reliance 2x2 on the FINAL diagnosis (needs gold_truth.csv).
  5. Failure chain by reflection band.
  6. IRR within each subject panel (blind P1 + P3 quality; clean same-condition).

Gold standard: optional gold_truth.csv (round_id,obj_index,truth 3/2/1/0).
AI per-objective stance: code_ai_stance() is a HEURISTIC stub — replace with a
human/LLM coding pass for publication-grade WOA / reliance numbers.
"""
import csv
import re
from collections import defaultdict, Counter
from pathlib import Path

import openpyxl
from yandaojie_rounds import build_rounds, DELIVERABLES

HERE = DELIVERABLES  # packets (in packets/), gold_truth.csv, and outputs live here
ROUNDS = {r["round_id"]: r for r in build_rounds()}

H = {
    "p1_diag": "【P1】盲诊_掌握度(3/2/1/0)", "p1_conf": "【P1】信心",
    "p2a_diag": "【P2a】看结论后_掌握度(3/2/1/0)", "p2a_conf": "【P2a】信心",
    "p2b_diag": "【P2b】再评_掌握度(3/2/1/0)", "p2b_conf": "【P2b】信心",
    "gnd": "【Q】诊断groundedness", "consist": "【Q】与盲诊(P1)一致性",
    "useful": "【Q】诊断有用性", "relevance": "【Q】追问相关性", "ontarget": "【Q】对准薄弱点",
    "elicit": "【Q】引出力", "fit": "【Q】适切", "attitude": "【Q】作答态度",
    "icap": "【Q】ICAP", "depth": "【Q】深度", "advance": "【Q】整轮推进",
    "teacher": "评分老师", "arm": "臂", "subject": "学科",
}


def packet_files():
    pdir = HERE / "packets"
    if pdir.is_dir():
        files = sorted(pdir.glob("packet_*.xlsx"))
        if files:
            return files
    return []


def parse_diag(text):
    """'目标1=2; 目标2=0' / '1:2,2:0' -> {0:2, 1:0}  (objective index 0-based)."""
    if not text:
        return {}
    out = {}
    for m in re.finditer(r"(?:目标)?\s*(\d+)\s*[=:：]\s*([0-3])", str(text)):
        out[int(m.group(1)) - 1] = int(m.group(2))
    return out


def to_int(v):
    try:
        return int(str(v).strip())
    except (ValueError, TypeError):
        return None


def load_records():
    recs = []
    for p in packet_files():
        ws = openpyxl.load_workbook(p, data_only=True)["评分表"]
        hdr = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        for r in range(2, ws.max_row + 1):
            def g(key):
                c = hdr.get(H[key])
                return ws.cell(r, c).value if c else None
            recs.append({
                "round_id": ws.cell(r, hdr["round_id"]).value,
                "subject": g("subject"), "teacher": g("teacher"), "arm": g("arm"),
                "p1": parse_diag(g("p1_diag")), "p2a": parse_diag(g("p2a_diag")),
                "p2b": parse_diag(g("p2b_diag")),
                "p1_conf": to_int(g("p1_conf")), "p2a_conf": to_int(g("p2a_conf")),
                "p2b_conf": to_int(g("p2b_conf")),
                "gnd": g("gnd"), "useful": to_int(g("useful")),
                "relevance": to_int(g("relevance")), "ontarget": to_int(g("ontarget")),
                "elicit": g("elicit"), "attitude": g("attitude"),
                "icap": g("icap"), "depth": to_int(g("depth")), "advance": to_int(g("advance")),
            })
    return recs


def code_ai_stance(round_id):
    """HEURISTIC AI per-objective stance on the 3/2/1/0 scale.
    targeted objective -> probed gap -> 1; else if AI listed any mastery -> 3; else 0.
    TODO: replace with a human/LLM coding pass for publication-grade WOA."""
    r = ROUNDS.get(round_id)
    if not r:
        return {}
    targeted = {t.get("objective_index") for t in (r["targeted_objectives"] or [])
                if isinstance(t.get("objective_index"), int)}
    has_mastery = bool((r["diagnoses"] or {}).get("mastered"))
    return {i: (1 if i in targeted else (3 if has_mastery else 0))
            for i in range(len(r["objectives"]))}


def load_gold():
    p = HERE / "gold_truth.csv"
    if not p.exists():
        return None
    gold = defaultdict(dict)
    with open(p, encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            gold[row["round_id"]][int(row["obj_index"])] = int(row["truth"])
    return gold


def fleiss_kappa(item_ratings):
    cats = sorted({c for it in item_ratings for c in it if c is not None})
    items = [[c for c in it if c is not None] for it in item_ratings]
    items = [it for it in items if len(it) >= 2]
    if not items or not cats:
        return None
    n = len(items)
    pj = {c: 0 for c in cats}
    Pi = []
    for it in items:
        m = len(it); cnt = Counter(it)
        for c in cats:
            pj[c] += cnt[c]
        Pi.append((sum(v * v for v in cnt.values()) - m) / (m * (m - 1)))
    total = sum(len(it) for it in items)
    for c in cats:
        pj[c] /= total
    Pe = sum(v * v for v in pj.values())
    return (sum(Pi) / n - Pe) / (1 - Pe) if (1 - Pe) else None


def krippendorff_alpha(units, metric="interval"):
    def delta(a, b):
        return (a - b) ** 2 if metric == "interval" else (0 if a == b else 1)
    units = [[v for v in u if v is not None] for u in units]
    units = [u for u in units if len(u) >= 2]
    if not units:
        return None
    coinc = defaultdict(float)
    marg = defaultdict(float)
    n = 0
    for u in units:
        m = len(u)
        for i in range(m):
            for j in range(m):
                if i != j:
                    coinc[(u[i], u[j])] += 1.0 / (m - 1)
        for v in u:
            marg[v] += 1
            n += 1
    if n < 2:
        return None
    vals = sorted(marg)
    Do = sum(coinc[(a, b)] * delta(a, b) for a in vals for b in vals)
    De = sum(marg[a] * marg[b] * delta(a, b) for a in vals for b in vals) / (n - 1)
    return 1 - Do / De if De else None


# ---------- influence helpers ----------
def infl(rows, frm, to, label):
    """Print + return switch/toward/WOA for a stage transition frm->to (keys in row)."""
    rows = [x for x in rows if x[frm] is not None and x[to] is not None]
    if not rows:
        return None
    changed = [x for x in rows if x[to] != x[frm]]
    tow = [x for x in changed if x["ai"] is not None and (x[to] - x[frm]) * (x["ai"] - x[frm]) > 0]
    woa = [(x[to] - x[frm]) / (x["ai"] - x[frm]) for x in rows
           if x["ai"] is not None and x["ai"] != x[frm]]
    switch = len(changed) / len(rows)
    toward = len(tow) / len(changed) if changed else None
    woa_m = sum(woa) / len(woa) if woa else None
    print(f"  [{label}] n={len(rows)}  改判率={switch:.1%}  "
          f"向AI移动(改判中)={'%.0f%%' % (100*toward) if toward is not None else '—'}  "
          f"WOA={'%.2f' % woa_m if woa_m is not None else '—'}")
    return {"n": len(rows), "switch": switch, "toward": toward, "woa": woa_m}


def conf_delta(records, a, b):
    vals = [r[b] - r[a] for r in records if r[a] is not None and r[b] is not None]
    return sum(vals) / len(vals) if vals else None


# ---------- main ----------
def main():
    recs = load_records()
    filled = [r for r in recs if r["p1"] and r["p2a"] and r["p2b"]]
    print(f"records loaded: {len(recs)} | with P1+P2a+P2b filled: {len(filled)}")
    if not filled:
        print("\n>>> No ratings yet. Skeleton OK — fill packets and re-run.")
        print(">>> Optional: add gold_truth.csv (round_id,obj_index,truth) for 2x2.")
        return

    gold = load_gold()

    inf = []  # per (record, objective)
    for r in filled:
        ai = code_ai_stance(r["round_id"])
        for oi in set(r["p1"]) & set(r["p2a"]) & set(r["p2b"]):
            inf.append({"rid": r["round_id"], "subject": r["subject"], "arm": r["arm"], "oi": oi,
                        "p1": r["p1"][oi], "p2a": r["p2a"][oi], "p2b": r["p2b"][oi],
                        "ai": ai.get(oi), "band": ROUNDS.get(r["round_id"], {}).get("band"),
                        "truth": (gold or {}).get(r["round_id"], {}).get(oi)})

    # 1. Advice effect (P1 -> P2a) — bare AI recommendation, all rounds
    print("\n=== 1. 建议效应 (P1→P2a, 仅给结论) ===")
    infl(inf, "p1", "p2a", "ALL")
    for subj in sorted({x["subject"] for x in inf if x["subject"]}):
        infl([x for x in inf if x["subject"] == subj], "p1", "p2a", subj)
    print(f"  ΔConf(P1→P2a) = {conf_delta(filled, 'p1_conf', 'p2a_conf'):+.2f}"
          if conf_delta(filled, 'p1_conf', 'p2a_conf') is not None else "  ΔConf: —")

    # 2. Staged increment (P2a -> P2b) by arm + NET reasoning effect (diff-in-diff)
    print("\n=== 2. 推理增量 (P2a→P2b) 按臂 + 净推理效应 ===")
    res = {}
    for armv in ("推理臂", "控制臂"):
        res[armv] = infl([x for x in inf if x["arm"] == armv], "p2a", "p2b", armv)
    if res.get("推理臂") and res.get("控制臂"):
        ns = res["推理臂"]["switch"] - res["控制臂"]["switch"]
        nw = ((res["推理臂"]["woa"] or 0) - (res["控制臂"]["woa"] or 0))
        print(f"  >>> 净推理效应 (推理臂−控制臂): 改判率 {ns:+.1%}   WOA {nw:+.2f}")
    cr = conf_delta([r for r in filled if r["arm"] == "推理臂"], "p2a_conf", "p2b_conf")
    cc = conf_delta([r for r in filled if r["arm"] == "控制臂"], "p2a_conf", "p2b_conf")
    if cr is not None and cc is not None:
        print(f"  ΔConf(P2a→P2b): 推理臂 {cr:+.2f}  控制臂 {cc:+.2f}  净 {cr-cc:+.2f}")

    # 3. Reliance 2x2 on FINAL diagnosis (P1 vs P2b), needs gold
    print("\n=== 3. 依赖 2x2 (最终判断 P1→P2b) ===")
    if not gold:
        print("  (skipped — provide gold_truth.csv)")
    else:
        cell = Counter()
        for x in inf:
            if x["truth"] is None or x["ai"] is None:
                continue
            moved = (x["p2b"] - x["p1"]) * (x["ai"] - x["p1"]) > 0 and x["p2b"] != x["p1"]
            ai_ok = x["ai"] == x["truth"]
            cell[("AI对" if ai_ok else "AI错", "改向AI" if moved else "未改向")] += 1
        for k, tag in [(("AI对", "改向AI"), "恰当依赖"), (("AI对", "未改向"), "算法厌恶"),
                       (("AI错", "改向AI"), "过度依赖(有害)"), (("AI错", "未改向"), "恰当自信")]:
            print(f"  {k[0]} × {k[1]}: {cell[k]:>4}  [{tag}]")

    # 4. Failure chain by reflection band
    print("\n=== 4. 失败链 by 反思厚度 ===")
    byband = defaultdict(lambda: {"n": 0, "halluc": 0, "offtgt": [], "diseng": 0, "woa": []})
    for r in filled:
        b = ROUNDS.get(r["round_id"], {}).get("band")
        d = byband[b]; d["n"] += 1
        if r["gnd"] in ("个别臆造", "多数臆造"):
            d["halluc"] += 1
        if r["relevance"] is not None:
            d["offtgt"].append(r["relevance"])
        if r["attitude"] in ("空答或乱码", "敷衍"):
            d["diseng"] += 1
    for x in inf:  # advice WOA (P1->P2a)
        if x["ai"] is not None and x["ai"] != x["p1"]:
            byband[x["band"]]["woa"].append((x["p2a"] - x["p1"]) / (x["ai"] - x["p1"]))
    for b in ["薄", "中", "厚"]:
        d = byband.get(b)
        if not d or not d["n"]:
            continue
        rel = sum(d["offtgt"]) / len(d["offtgt"]) if d["offtgt"] else float("nan")
        woa = sum(d["woa"]) / len(d["woa"]) if d["woa"] else float("nan")
        print(f"  {b}: n={d['n']:>3} 臆造率={d['halluc']/d['n']:.0%} "
              f"追问相关均分={rel:.2f} 脱离率={d['diseng']/d['n']:.0%} 建议WOA={woa:.2f}")
    _plot_failure_chain(byband)

    # 5. IRR within each subject panel (blind P1 + P3 quality = same-condition for all 3)
    print("\n=== 5. 评分员信度 (IRR) — 各学科面板 ===")
    def by_round(rows, field):
        m = defaultdict(list)
        for r in rows:
            m[r["round_id"]].append(r[field])
        return [v for v in m.values() if len([x for x in v if x is not None]) >= 2]

    def p1_by_round(rows):
        m = defaultdict(list)
        for r in rows:  # mean blind mastery per record, then agreement across teachers
            if r["p1"]:
                m[r["round_id"]].append(round(sum(r["p1"].values()) / len(r["p1"])))
        return [v for v in m.values() if len(v) >= 2]

    def irr(rows, label):
        if not rows:
            return
        f = lambda v: f"{v:.2f}" if isinstance(v, float) else str(v)
        print(f"  [{label:>4}] 盲诊P1 α={f(krippendorff_alpha(p1_by_round(rows)))}  "
              f"groundedness κ={f(fleiss_kappa(by_round(rows, 'gnd')))}  "
              f"ICAP κ={f(fleiss_kappa(by_round(rows, 'icap')))}  "
              f"追问相关 α={f(krippendorff_alpha([[to_int(x) for x in v] for v in by_round(rows, 'relevance')]))}")

    for subj in sorted({r["subject"] for r in filled if r["subject"]}):
        irr([r for r in filled if r["subject"] == subj], subj)
    irr(filled, "pooled")


def _plot_failure_chain(byband):
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except Exception:
        print("  (matplotlib not available — skipped figure)")
        return
    bands = [b for b in ["薄", "中", "厚"] if byband.get(b, {}).get("n")]
    if not bands:
        return
    xs = [{"薄": "thin", "中": "mid", "厚": "thick"}[b] for b in bands]
    hall = [byband[b]["halluc"] / byband[b]["n"] for b in bands]
    woa = [sum(byband[b]["woa"]) / len(byband[b]["woa"]) if byband[b]["woa"] else 0 for b in bands]
    fig, ax1 = plt.subplots(figsize=(5, 3.2))
    ax1.plot(xs, hall, "o-", color="#C0392B", label="hallucination rate")
    ax1.set_ylabel("AI hallucination rate", color="#C0392B")
    ax2 = ax1.twinx()
    ax2.plot(xs, woa, "s--", color="#2E5496", label="advice WOA")
    ax2.set_ylabel("advice WOA (expert persuaded)", color="#2E5496")
    ax1.set_xlabel("student reflection thickness")
    fig.suptitle("Double jeopardy: thinner reflection -> more hallucination & more persuasion")
    fig.tight_layout()
    fig.savefig(HERE / "failure_chain.png", dpi=130)
    print("  -> saved failure_chain.png")


if __name__ == "__main__":
    main()
