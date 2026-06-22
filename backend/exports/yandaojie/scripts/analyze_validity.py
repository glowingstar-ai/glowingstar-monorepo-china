"""Analyze the filled defense-depth annotation packets -> the validity-gap numbers.

Computes: over-credit rate (+95% CI), divergence rate, defense-depth distribution,
cross-subject breakdown, and inter-rater reliability (Krippendorff alpha / Fleiss
kappa) per subject panel. Runs on empty packets (reports awaiting); falls back to
the LLM-judge preliminary (_validity_ratings.json) so you can see the pipeline +
the number to beat.
"""
import csv, json, math
from collections import defaultdict, Counter
from pathlib import Path
import openpyxl
from yandaojie_rounds import DELIVERABLES

PDIR = DELIVERABLES / "packets_validity"
MC_MAP = {"已掌握": "got_it", "部分": "partial", "未掌握": "missed"}
DEPTH_MAP = {"扎实": "sound", "对标签错机制": "label_wrong_mech", "空洞或猜对": "hollow",
             "完全不懂": "none", "难判断": "cant_tell"}
NOT_SOUND = {"label_wrong_mech", "hollow", "none"}


def ci(p, n):
    e = 1.96 * math.sqrt(p * (1 - p) / n) if n else 0
    return (round(100 * (p - e)), round(100 * (p + e)))


def fleiss_kappa(items):
    items = [[c for c in it if c is not None] for it in items]
    items = [it for it in items if len(it) >= 2]
    cats = sorted({c for it in items for c in it})
    if not items or not cats:
        return None
    n = len(items); pj = {c: 0 for c in cats}; Pi = []
    for it in items:
        m = len(it); cnt = Counter(it)
        for c in cats:
            pj[c] += cnt[c]
        Pi.append((sum(v * v for v in cnt.values()) - m) / (m * (m - 1)))
    tot = sum(len(it) for it in items)
    for c in cats:
        pj[c] /= tot
    Pe = sum(v * v for v in pj.values())
    return (sum(Pi) / n - Pe) / (1 - Pe) if (1 - Pe) else None


def kripp_alpha(units, metric="nominal", order=None):
    def num(v): return order.index(v) if order else v
    def delta(a, b):
        if metric == "ordinal":
            return (num(a) - num(b)) ** 2
        return 0 if a == b else 1
    units = [[v for v in u if v is not None] for u in units]
    units = [u for u in units if len(u) >= 2]
    if not units:
        return None
    co = defaultdict(float); marg = defaultdict(float); n = 0
    for u in units:
        m = len(u)
        for i in range(m):
            for j in range(m):
                if i != j:
                    co[(u[i], u[j])] += 1.0 / (m - 1)
        for v in u:
            marg[v] += 1; n += 1
    if n < 2:
        return None
    vals = sorted(marg, key=lambda x: (order.index(x) if order else 0, str(x)))
    Do = sum(co[(a, b)] * delta(a, b) for a in vals for b in vals)
    De = sum(marg[a] * marg[b] * delta(a, b) for a in vals for b in vals) / (n - 1)
    return 1 - Do / De if De else None


def load_packets():
    recs = []
    if not PDIR.is_dir():
        return recs
    for f in sorted(PDIR.glob("packet_*.xlsx")):
        parts = f.stem.split("_")  # packet_<subject>_T#
        subject, teacher = parts[1], parts[2]
        ws = openpyxl.load_workbook(f, data_only=True)["标注表"]
        hdr = {(ws.cell(1, c).value or ""): c for c in range(1, ws.max_column + 1)}
        def col(prefix):
            for h, c in hdr.items():
                if h.startswith(prefix):
                    return c
            return None
        c_ep, c_mc, c_dep = col("序号"), col("①MC视角"), col("②辩护深度")
        for r in range(2, ws.max_row + 1):
            mc = MC_MAP.get((ws.cell(r, c_mc).value or "").strip()) if c_mc else None
            dep = DEPTH_MAP.get((ws.cell(r, c_dep).value or "").strip()) if c_dep else None
            if mc or dep:
                recs.append({"epid": ws.cell(r, c_ep).value, "subject": subject,
                             "teacher": teacher, "mc": mc, "depth": dep})
    return recs


def report(records, label):
    rated = [r for r in records if r["mc"] and r["depth"]]
    print(f"\n===== {label} =====")
    print(f"ratings with mc+depth: {len(rated)}")
    if not rated:
        return
    over = [r for r in rated if r["mc"] == "got_it" and r["depth"] in NOT_SOUND]
    gotit = [r for r in rated if r["mc"] == "got_it"]
    p_over = len(over) / len(rated)
    print(f"过度给分率(全样本): {100*p_over:.0f}% (n={len(rated)}, 95%CI {ci(p_over,len(rated))})")
    if gotit:
        p_div = len(over) / len(gotit)
        print(f"分歧率(MC判过中): {100*p_div:.0f}% ({len(over)}/{len(gotit)}, CI {ci(p_div,len(gotit))})")
    dd = Counter(r["depth"] for r in rated)
    lw = dd["label_wrong_mech"] / len(rated)
    print(f"对标签错机制: {100*lw:.0f}% (CI {ci(lw,len(rated))})  深度分布: {dict(dd)}")
    print("按学科 过度给分率:")
    bysub = defaultdict(list)
    for r in rated:
        bysub[r["subject"]].append(r)
    for sub, rs in sorted(bysub.items()):
        o = sum(1 for r in rs if r["mc"] == "got_it" and r["depth"] in NOT_SOUND)
        print(f"  {sub}: {100*o/len(rs):.0f}% ({o}/{len(rs)})")


def reliability(records):
    print("\n===== 评分员信度 (IRR) — 各学科 panel =====")
    bysub = defaultdict(lambda: defaultdict(dict))  # subject -> epid -> teacher -> rec
    for r in records:
        bysub[r["subject"]][r["epid"]][r["teacher"]] = r
    fmt = lambda v: f"{v:.2f}" if isinstance(v, float) else str(v)
    for sub in ["数学", "科学", "英语"]:
        eps = bysub.get(sub)
        if not eps:
            continue
        mc_units, dep_units, over_units = [], [], []
        for ep, by_t in eps.items():
            mcs = [r["mc"] for r in by_t.values() if r["mc"]]
            deps = [r["depth"] for r in by_t.values() if r["depth"]]
            ovs = [(r["mc"] == "got_it" and r["depth"] in NOT_SOUND) for r in by_t.values() if r["mc"] and r["depth"]]
            if len(mcs) >= 2: mc_units.append(mcs)
            if len(deps) >= 2: dep_units.append(deps)
            if len(ovs) >= 2: over_units.append(ovs)
        a_mc = kripp_alpha(mc_units, "ordinal", order=["missed", "partial", "got_it"])
        a_dep = kripp_alpha(dep_units, "nominal")
        k_over = fleiss_kappa(over_units)
        print(f"  [{sub}] MC视角 α={fmt(a_mc)}  辩护深度 α={fmt(a_dep)}  过度给分 κ={fmt(k_over)}  (n_ep={len(mc_units)})")


def main():
    recs = load_packets()
    print(f"loaded {len(recs)} annotations from {PDIR.name}/")
    if any(r["mc"] and r["depth"] for r in recs):
        report(recs, "教师标注 (TEACHER)")
        reliability(recs)
    else:
        print(">>> 老师还没填。骨架就绪——填完 packets_validity/ 后重跑。")
        # fallback: show the LLM-judge preliminary (the number to beat)
        f = DELIVERABLES / "_validity_ratings.json"
        if f.exists():
            raw = json.load(open(f))
            eps = {e["idx"]: e for e in json.load(open(DELIVERABLES / "_validity_episodes.json"))}
            llm = [{"epid": r["idx"], "subject": eps.get(r["idx"], {}).get("subject", "?"),
                    "teacher": "LLM", "mc": r["mc_view"],
                    "depth": {"sound": "sound", "label_right_mechanism_wrong": "label_wrong_mech",
                              "hollow_or_guess": "hollow", "no_understanding": "none", "cant_tell": "cant_tell"}.get(r["defense_depth"])}
                   for r in raw]
            report(llm, "LLM-judge 预研 (n=106, 单评分者, 待教师验证)")


if __name__ == "__main__":
    main()
