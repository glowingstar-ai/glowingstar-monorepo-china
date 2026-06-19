"""Generate 3 per-teacher rating packets from the 140-round sample.

Design:
* Every round is rated by all 3 teachers (full crossing -> inter-rater reliability).
* ±reasoning is a WITHIN-ITEM crossover: for each round the 3 teachers are split
  across conditions via a 6-row balanced rotation, so (a) every round is seen in
  BOTH conditions and (b) each teacher sees ~50% +推理 / ~50% 仅结论.
* Each packet presents the 140 rounds in a teacher-specific randomized order
  (deterministic md5), so order effects don't align across teachers.

Outputs: packet_T1.xlsx, packet_T2.xlsx, packet_T3.xlsx, assignment_key.csv
"""
import csv
import hashlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from yandaojie_rounds import sampled_rounds, DELIVERABLES

TEACHERS = ["T1", "T2", "T3"]
SUBJECTS = ["数学", "科学", "英语"]  # each subject gets its own 3 subject-matter teachers
# 6-row balanced rotation of (T1,T2,T3) conditions; R=+推理, C=仅结论
ROT = [
    ("+推理", "+推理", "仅结论"),
    ("+推理", "仅结论", "+推理"),
    ("仅结论", "+推理", "+推理"),
    ("仅结论", "仅结论", "+推理"),
    ("仅结论", "+推理", "仅结论"),
    ("+推理", "仅结论", "仅结论"),
]

FONT = "Arial"
FILLS = {
    "id": PatternFill("solid", fgColor="DDE6F2"), "ctx": PatternFill("solid", fgColor="EEF3FA"),
    "p1": PatternFill("solid", fgColor="E7F4E4"), "p2": PatternFill("solid", fgColor="CDE7C8"),
    "q": PatternFill("solid", fgColor="FFF6D5"), "meta": PatternFill("solid", fgColor="F0F0F0"),
}
HEAD = {"ctx": "1F3864", "id": "1F3864", "p1": "2E7D32", "p2": "1B5E20", "q": "B8860B", "meta": "555555"}
thin = Side(style="thin", color="C9C9C9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)

# (header, round-key or None, width, dropdown-kind, fill)
COLS = [
    ("round_id", "round_id", 11, None, "id"),
    ("学生ID", "student_id", 7, None, "ctx"),
    ("学科", "subject", 7, None, "ctx"),
    ("课题", "topic", 15, None, "ctx"),
    ("轮次", "round", 5, None, "ctx"),
    ("本课学习目标", "objectives_text", 34, None, "ctx"),
    ("学生反思（输入）", "reflection_text", 30, None, "ctx"),
    ("上一轮·AI问题", "prev_q", 20, None, "ctx"),
    ("上一轮·学生答", "prev_a", 16, None, "ctx"),
    ("【P1】盲诊_掌握度(3/2/1/0)", None, 16, None, "p1"),
    ("【P1】我会怎么问", None, 20, None, "p1"),
    ("【P1】信心", None, 7, "S15", "p1"),
    ("AI诊断·已掌握", "ai_mastered_text", 26, None, "ctx"),
    ("AI诊断·未掌握", "ai_not_mastered_text", 26, None, "ctx"),
    ("AI针对目标(理由)", "ai_targets_text", 24, None, "ctx"),
    ("AI推理过程〔仅+推理条件〕", "__reasoning_cond__", 40, None, "ctx"),
    ("【P2】重评_掌握度(3/2/1/0)", None, 16, None, "p2"),
    ("【P2】信心", None, 7, "S15", "p2"),
    ("【P2】改判说明(可选)", None, 20, None, "p2"),
    ("【Q】诊断groundedness", None, 14, "DIAG_GND", "q"),
    ("【Q】臆造说明", None, 16, None, "q"),
    ("【Q】与盲诊(P1)一致性", None, 12, "CONSIST", "q"),
    ("【Q】诊断有用性", None, 8, "S15", "q"),
    ("AI追问", "ai_question", 30, None, "ctx"),
    ("【Q】追问相关性", None, 8, "S15", "q"),
    ("【Q】对准薄弱点", None, 9, "S15", "q"),
    ("【Q】引出力", None, 9, "ELICIT", "q"),
    ("【Q】适切", None, 7, "S15", "q"),
    ("【Q】会用吗", None, 8, "YESNO", "q"),
    ("【Q】会用原因", None, 14, None, "q"),
    ("学生回答", "student_answer", 24, None, "ctx"),
    ("【Q】作答态度", None, 12, "ATTITUDE", "q"),
    ("【Q】ICAP", None, 10, "ICAP", "q"),
    ("【Q】深度", None, 7, "S15", "q"),
    ("【Q】正确性", None, 11, "CORRECT", "q"),
    ("【Q】整轮推进", None, 9, "S15", "q"),
    ("【Q】匹配度", None, 8, "S15", "q"),
    ("评分老师", "__teacher__", 9, None, "meta"),
    ("条件", "__cond__", 9, None, "meta"),
    ("备注", None, 16, None, "meta"),
]
DV = {"S15": '"1,2,3,4,5"', "DIAG_GND": '"全部有据,个别臆造,多数臆造"', "CONSIST": '"一致,部分一致,矛盾"',
      "ELICIT": '"L1回忆,L2解释,L3论证,L4迁移"', "YESNO": '"是,否"', "ATTITUDE": '"真诚尝试,敷衍,空答或乱码"',
      "ICAP": '"被动,主动,建构,互动"', "CORRECT": '"正确,部分正确,错误"'}
RCOLS = [
    ("round_id", "round_id", 11, None, "id"), ("学科", "subject", 7, None, "ctx"),
    ("课题", "topic", 14, None, "ctx"), ("轮次", "round", 5, None, "ctx"),
    ("学生反思（输入）", "reflection_text", 28, None, "ctx"),
    ("AI诊断·已掌握", "ai_mastered_text", 24, None, "ctx"),
    ("AI诊断·未掌握", "ai_not_mastered_text", 24, None, "ctx"),
    ("AI追问", "ai_question", 26, None, "ctx"), ("AI推理过程", "reasoning", 52, None, "ctx"),
    ("⑥a推理逻辑", None, 9, "S15", "p2"), ("⑥b faithfulness", None, 11, "S15", "p2"),
    ("⑥c有无捏造", None, 10, "FAB", "p2"), ("评分老师", "__teacher__", 9, None, "meta"),
    ("备注", None, 16, None, "meta"),
]


def order_for(teacher, rounds):
    return sorted(rounds, key=lambda r: hashlib.md5(f"{r['round_id']}|{teacher}".encode()).hexdigest())


def write_sheet(ws, cols, rounds, teacher, cond_map, dvmap):
    ws.sheet_view.showGridLines = False
    for ci, (hdr, key, width, dv, fg) in enumerate(cols, start=1):
        c = ws.cell(1, ci, hdr)
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=HEAD[fg]); c.alignment = CTR; c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 46
    for ri, r in enumerate(rounds, start=2):
        cond = cond_map[r["round_id"]]
        for ci, (hdr, key, width, dv, fg) in enumerate(cols, start=1):
            c = ws.cell(ri, ci); c.border = BORDER; c.font = Font(name=FONT, size=10)
            if key == "__reasoning_cond__":
                c.value = r["reasoning"] if cond == "+推理" else "〔仅结论条件：本行不提供推理〕"
                c.fill = FILLS["ctx"]; c.alignment = WRAP
            elif key == "__teacher__":
                c.value = teacher; c.fill = FILLS["meta"]; c.alignment = CTR
            elif key == "__cond__":
                c.value = cond; c.fill = FILLS["meta"]; c.alignment = CTR
            elif key is not None:
                c.value = r.get(key, ""); c.fill = FILLS["id"] if fg == "id" else FILLS[fg]
                c.alignment = CTR if fg == "id" else WRAP
            else:
                c.fill = FILLS[fg]; c.alignment = CTR if dv in dvmap else WRAP
        ws.row_dimensions[ri].height = 96
    ws.freeze_panes = "B2"
    nrow = len(rounds) + 1
    for kind, formula in dvmap.items():
        d = DataValidation(type="list", formula1=formula, allow_blank=True, showErrorMessage=False)
        ws.add_data_validation(d)
        for ci, (hdr, key, width, k, fg) in enumerate(cols, start=1):
            if k == kind:
                L = get_column_letter(ci); d.add(f"{L}2:{L}{nrow}")


def main():
    packdir = DELIVERABLES / "packets"
    packdir.mkdir(parents=True, exist_ok=True)
    sample = sampled_rounds()  # 140, canonical order (grouped by subject)
    key_rows = []

    for subj in SUBJECTS:
        srs = [r for r in sample if r["subject"] == subj]  # canonical order preserved
        # ±reasoning crossover assigned WITHIN this subject's 3 teachers
        cond = {t: {} for t in TEACHERS}
        for i, r in enumerate(srs):
            pat = ROT[i % len(ROT)]
            for t, cv in zip(TEACHERS, pat):
                cond[t][r["round_id"]] = cv

        for t in TEACHERS:
            wb = Workbook()
            ws = wb.active; ws.title = "评分表"
            ordered = order_for(f"{subj}|{t}", srs)  # subject-specific row order
            write_sheet(ws, COLS, ordered, t, cond[t], DV)
            ws2 = wb.create_sheet("推理审计")
            plus = [r for r in ordered if cond[t][r["round_id"]] == "+推理"]
            write_sheet(ws2, RCOLS, plus, t, cond[t], {"S15": '"1,2,3,4,5"', "FAB": '"有,无"'})
            wb.save(str(packdir / f"packet_{subj}_{t}.xlsx"))
            n = sum(1 for v in cond[t].values() if v == "+推理")
            print(f"{subj} {t}: +推理 {n}/{len(srs)}  仅结论 {len(srs)-n}")

        mix = sum(1 for r in srs if len({cond[t][r['round_id']] for t in TEACHERS}) == 2)
        print(f"  {subj}: rounds seen in BOTH conditions: {mix}/{len(srs)}")
        for i, r in enumerate(srs):
            key_rows.append([subj, i, r["round_id"], r["band"], r["stratum"],
                             cond["T1"][r["round_id"]], cond["T2"][r["round_id"]], cond["T3"][r["round_id"]]])

    with open(DELIVERABLES / "assignment_key.csv", "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        w.writerow(["subject", "idx_in_subject", "round_id", "band", "stratum",
                    "T1_cond", "T2_cond", "T3_cond"])
        w.writerows(key_rows)
    print(f"wrote {len(SUBJECTS) * len(TEACHERS)} packets to deliverables/packets/ + assignment_key.csv")


if __name__ == "__main__":
    main()
